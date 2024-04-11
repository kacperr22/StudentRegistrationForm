from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox

import filetype as filetype
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#89d6fb"  # background
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No"
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Date of Birth"
    sheet['F1'] = "Religion"
    sheet['G1'] = "Date of registration"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father name"
    sheet['J1'] = "Mother name"
    sheet['K1'] = "Occupation"
    sheet['L1'] = "Occupation"

    file.save('Student_data.xlsx')
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

def search():
    text = Search.get() # taking input form entry box
    Clear() # to clear all the data already available in entry box and other
    saveButton.config(state='disable') # after clicking on search, save button will disable

    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1] # show the position like A2, A3 ... n
            reg_number = str(name)[15:-1] # show the number like 2, 3 ... n

    try:
        pass
    except:
        messagebox.showerror('Invalid', 'Invalid registration number!')

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    Registation.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4 == 'Famale':
        R2.select()
    else:
        R1.select()

    DateOfBirth.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    fatherName.set(x9)
    motherName.set(x10)
    fatherOccumpation.set(x11)
    motherOccumpation.set(x12)

    img = (Image.open("Student_images/"+str(x1)+".jpg"))
    resized_imaged = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_imaged)
    lbl.config(image=photo2)
    lbl.image=photo2

def Update():
    R1 = Registation.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DateOfBirth.get()
    D1 = Date.get()
    RE1 = Religion.get()
    S1 = Skill.get()
    FN = fatherName.get()
    MN = motherName.get()
    FO = fatherOccumpation.get()
    MO = motherOccumpation.get()

    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            reg_no_position=str(name)[14:-1]
            reg_number = str(name)[15:-1]

    ## sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=N1)
    sheet.cell(column=3, row=int(reg_number), value=C1)
    sheet.cell(column=4, row=int(reg_number), value=G1)
    sheet.cell(column=5, row=int(reg_number), value=D2)
    sheet.cell(column=6, row=int(reg_number), value=D1)
    sheet.cell(column=7, row=int(reg_number), value=RE1)
    sheet.cell(column=8, row=int(reg_number), value=S1)
    sheet.cell(column=9, row=int(reg_number), value=FN)
    sheet.cell(column=10, row=int(reg_number), value=MN)
    sheet.cell(column=11, row=int(reg_number), value=FO)
    sheet.cell(column=12, row=int(reg_number), value=MO)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Student_images/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo('Update', 'Updated succesfully')
    Clear()

def Exit():
    root.destroy()


# function add student's photo to registration system
def show_image():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), \
                                          title='Select image file',
                                          filetypes=(("JPG File", "*.jpg"), ("PNG", "*.png"), ("All files", "*.*")))

    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

# function created to automatic enter registration number
def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registation.set(max_row_value+1)
    except:
        Registation.set("1")

def Clear():
    Name.set(' ')
    DateOfBirth.set(' ')
    Religion.set(' ')
    Skill.set(' ')
    fatherName.set(' ')
    motherName.set(' ')
    fatherOccumpation.set(' ')
    motherOccumpation.set(' ')
    Class.set('Select Class')

    registration_no()

    saveButton.config(state='normal')
    img1 = PhotoImage(file='Images/upload photo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=''

def Save():
    R1 = Registation.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror('ERROR', 'SELECT GENDER!')

    D2=DateOfBirth.get()
    D1 = Date.get()
    RE1 = Religion.get()
    S1 = Skill.get()
    FN = fatherName.get()
    MN = motherName.get()
    FO = fatherOccumpation.get()
    MO = motherOccumpation.get()

    if N1 =='' or C1=='Select Class' or D2 =='' or RE1=='' or S1=='' or FN=='' or MN=='' or FO=='' or MO=='':
        messagebox.showerror('ERROR', 'FEW DATA IS MISSING.')
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=RE1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=FN)
        sheet.cell(column=10, row=sheet.max_row, value=MN)
        sheet.cell(column=11, row=sheet.max_row, value=FO)
        sheet.cell(column=12, row=sheet.max_row, value=MO)
        file.save(r'Student_data.xlsx')

        try:
            img.save("Student_images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("INFO", "Profile Picture is not available")

        messagebox.showinfo("INFO", "Succesfully data entered!!!")

        Clear() # Clear entry box and image section

        registration_no()



# create top frames
Label(root, text="Email: kacper.kluge@gmail.com", width=10, height=2, bg="white", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#01303f", fg="#fff",
      font='Arial 20 bold').pack(side=TOP, fill=X)

# create search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font='arail 20').place(x=820, y=50)
imageicon3 = PhotoImage(file="Images/search.png")
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg='#02a9f7', font='arial 12 bold', command=search)
Srch.place(x=1060, y=48)

imageicon4 = PhotoImage(file="Images/Layer 4.png")
Update_button = Button(root, image=imageicon4, bg='#02a9f7', command=Update)
Update_button.place(x=110, y=46)

# Registation and date
Label(root, text="Registartion No:", font="arial 13", fg=framebg, bg=background).place(x=30, y=140)
Label(root, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=140)

Registation = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registation, width=15, font='arial 10')
reg_entry.place(x=160, y=140)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font='arial 10')
date_entry.place(x=555, y=140)

Date.set(d1)

# Student details
labelframe1 = LabelFrame(root, text="Student details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250,
                         relief=GROOVE)
labelframe1.place(x=30, y=190)

Label(labelframe1, text="Full Name: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=30, y=50)
Label(labelframe1, text="Date of Birth: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=30, y=100)
Label(labelframe1, text="Gender: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=30, y=150)

Label(labelframe1, text="Class: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=500, y=50)
Label(labelframe1, text="Religion: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=500, y=100)
Label(labelframe1, text="Skills: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(labelframe1, textvariable=Name, width=20, font='arail 10')
name_entry.place(x=150, y=50)

radio = IntVar()
R1 = Radiobutton(labelframe1, text='Male', variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(labelframe1, text='Female', variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

DateOfBirth = StringVar()
dateofbirth_entry = Entry(labelframe1, textvariable=DateOfBirth, width=20, font='arail 10')
dateofbirth_entry.place(x=150, y=100)

Religion = StringVar()
religion_entry = Entry(labelframe1, textvariable=Religion, width=20, font='arail 10')
religion_entry.place(x=620, y=100)

Skill = StringVar()
skill_entry = Entry(labelframe1, textvariable=Skill, width=20, font='arail 10')
skill_entry.place(x=620, y=150)

Class = Combobox(labelframe1, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font='Roboto 10',
                 width=17, state='r')
Class.place(x=620, y=50)
Class.set("Select Class")

# Parent's details
labelframe2 = LabelFrame(root, text="Parent's details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220,
                         relief=GROOVE)
labelframe2.place(x=30, y=460)

Label(labelframe2, text="Father's Name: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=30, y=50)
Label(labelframe2, text="Occupation: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=30, y=100)

fatherName = StringVar()
fatherName_entry = Entry(labelframe2, textvariable=fatherName, width=20, font='arial 10')
fatherName_entry.place(x=160, y=50)

fatherOccumpation = StringVar()
fatherOccupation_entry = Entry(labelframe2, textvariable=fatherOccumpation, width=20, font='arial 10')
fatherOccupation_entry.place(x=160, y=100)

Label(labelframe2, text="Mother's Name: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=500, y=50)
Label(labelframe2, text="Occupation: ", font='arial 13 bold', bg=framebg, fg=framefg).place(x=500, y=100)

motherName = StringVar()
motherName_entry = Entry(labelframe2, textvariable=motherName, width=20, font='arial 10')
motherName_entry.place(x=630, y=50)

motherOccumpation = StringVar()
motherOccupation_entry = Entry(labelframe2, textvariable=motherOccumpation, width=20, font='arial 10')
motherOccupation_entry.place(x=630, y=100)

# image
imageFrame = Frame(root, bd=3, bg='black', width=200, height=200, relief=GROOVE)
imageFrame.place(x=1000, y=150)

img = PhotoImage(file='Images/upload photo.png')
lbl = Label(imageFrame, image=img)
lbl.place(x=0, y=0)

# buttons

uploadButton = Button(root, text='Upload', width=19, height=2, font='arial 12 bold', bg='lightblue', command=show_image)
uploadButton.place(x=1000, y=370)

saveButton = Button(root, text='Save', width=19, height=2, font='arial 12 bold', bg='lightgreen', command=Save)
saveButton.place(x=1000, y=450)

resetButton = Button(root, text='Reset', width=19, height=2, font='arial 12 bold', bg='grey', command=Clear)
resetButton.place(x=1000, y=530)

exitButton = Button(root, text='Exit', width=19, height=2, font='arial 12 bold', bg='red', command=Exit)
exitButton.place(x=1000, y=610)

root.mainloop()
